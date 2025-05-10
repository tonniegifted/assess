from getpass import win_getpass
# from db import dbfunc
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
import sys
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from datetime import datetime
from fpdf import FPDF
import os




# Step 1: Enable High DPI scaling (must be before app is created)
QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

# Step 2: Connect to DB (unchanged)
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="print",
    database="assess",
    buffered=True
)
cur = db.cursor()

# Get absolute path for resource files (for PyInstaller)
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# Step 3: Main window class
class mainwin(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("home.ui"), self)
        self.loadschoolname()
        self.loadterm()
        self.enterscoresbutton.clicked.connect(self.loadenterscore)
        self.exitbutton.clicked.connect(self.close_application)
        self.updatebutton.clicked.connect(self.loadupdatescore)
        self.deletebutton.clicked.connect(self.loaddeletescore)
        self.adminbutton.clicked.connect(self.loadadminpanel)
        self.analysisbutton.clicked.connect(self.loadanalysis)
        self.managelearner.clicked.connect(self.loadsavelearner)
        self.setschool()
        
        
    def setschool(self):
        school_path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
        with open(school_path,"r") as file:
            school=file.read()
            self.level_label.setText(school)

    def loadschoolname(self):
        path="D:/TONNIEGIFTED/Documents/programs/Remedial2/name.txt"
        with open(path,"r") as file:
            schoolname=file.read()
            self.schoolnamelabel.setText(schoolname)

    def loadterm(self):
            cur.execute("""SELECT selected_term,selected_year
                        FROM term WHERE is_active=1""")
            term=cur.fetchone()
            term_year=f"Term {term[0]}, {term[1]}"
            self.termlabel.setText(term_year) 

    #Loading windows/screens
    def loadenterscore(self):
        enterscorewin=enterscore()
        widget.addWidget(enterscorewin)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def loadsavelearner(self):
        savelearnerwin=savelearner()
        widget.addWidget(savelearnerwin)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def loadupdatescore(self):
        updatescorewin=updatescore()
        widget.addWidget(updatescorewin)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
    def loaddeletescore(self):
        deletescorewin=deletescore()
        widget.addWidget(deletescorewin)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def loadadminpanel(self):
        adminwin=adminpanel()
        widget.addWidget(adminwin)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    def loadanalysis(self):
        analysiswin=analysis()
        widget.addWidget(analysiswin)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def close_application(self):
        """Exit confirmation with system bell and complete window cleanup"""
        # Play system alert sound
        QApplication.beep()

        # Show confirmation dialog
        reply = QMessageBox.question(
            self,
            'Confirm Exit',
            'Are you sure you want to\nexit the program?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Close all application windows
            for window in QApplication.topLevelWidgets():
                window.close()

            # Ensure complete application termination
            QApplication.quit()


class enterscore(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("enterscores.ui"), self)

        # Initialize grade and subject
        
        
        
        self.readschool()

        # Connect signals
        self.gradecombo.currentTextChanged.connect(self.update_grade_subject)
        self.gradecombo.currentTextChanged.connect(self.loadlearners)
        self.gradecombo.currentTextChanged.connect(self.displayentered)
        self.totalbutton.clicked.connect(self.savetotalscore)
        self.subjectcombo.currentTextChanged.connect(self.update_grade_subject)
        self.subjectcombo.currentTextChanged.connect(self.hidetotalscore)
        self.enterscoresbutton.clicked.connect(self.savescores)
        self.homebutton.clicked.connect(self.homepage)
    #display grades     
    def readschool(self):
        school_path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
        with open(school_path,"r") as file:
            school=file.read()
            # self.level_label.setText(school)
        if school =="Junior School":
            grades=['Seven','Eight','Nine']
            subjects=['MATHS','ENG','KISW','INT','SST','PTC','AGN','CRE','CAS']
        elif school=="Upper Primary":
            grades=['Four','Five','Six']
            subjects=['MATHS','ENG','KISW','SCIE','SST','AGN','CRE','C/A']
        else:
            grades=['One','Two','Three']
            subjects=['MA/ACT','EN/ACT','KI/ACT','ENV/ACT','RE/ACT','C/ACT']
            
        self.gradecombo.addItems(grades)
        self.subjectcombo.addItems(subjects)
        self.update_grade_subject()
        self.displayentered()
        self.loadlearners(self.grade)
        self.hidetotalscore()
        self.displayentered()
        # Load initial data
       
        
    def displayentered(self):
        try:
            self.enteredscores.clear()
            # Get grade_id from combobox
            cur.execute("""SELECT grade_id FROM grade WHERE grade_name = %s""",
                    (self.gradecombo.currentText(),)) 
            grade_id = cur.fetchone()[0]
            
            # Get active term_id
            cur.execute("""SELECT term_id FROM term WHERE is_active = 1""")
            term_id = cur.fetchone()[0]
            
            # Determine school level and set appropriate subject_ids
            school_path = "D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
            with open(school_path, "r") as file:
                school = file.read().strip()
            
            if school == 'Lower Primary':
                subject_ids = [(10,), (11,), (12,), (13,), (14,), (15,)]
                required_subjects = {"KI/ACT", "EN/ACT", "MA/ACT", "RE/ACT", "ENV/ACT", "C/ACT"}
            elif school == 'Upper Primary':
                subject_ids = [(1,), (2,), (3,), (5,), (6,), (7,), (16,), (17,)]
                required_subjects = {"MATHS", "ENG", "KISW", "SST", "AGN", "CRE", "SCIE", "C/A"}
            elif school == 'Junior School':
                subject_ids = [(1,), (2,), (3,), (4,), (5,), (6,), (7,), (8,), (9,)]
                required_subjects = {"MATHS", "ENG", "KISW", "INT", "SST", "AGN", "CRE", "CAS", "PTC"}
            else:
                subject_ids = []
                required_subjects = set()
            
            output_text = ""

            for subject_id in subject_ids:
                cur.execute("""
                    SELECT s.subject_score, a.subject_abbr FROM score s
                    JOIN subject a ON a.subject_id = s.subject_id
                    WHERE s.subject_id = %s AND s.grade_id = %s AND s.term_id = %s 
                    LIMIT 1
                """, (subject_id[0], grade_id, term_id))
                
                result = cur.fetchone()
                if result:
                    output_text += f"{result[1]} | "

            output_text = output_text.rstrip(" | ")
            self.enteredscores.setText(output_text)

            # Add the readyforanalysis check with school-specific required subjects
            if all(subject in output_text for subject in required_subjects):
                self.enteredscores.clear()
                self.enteredscores.setText(f"Grade scores ready for analysis")
                self.enteredscores.setStyleSheet("color:red; font-size:9px")
        except Exception as e:
            QMessageBox.critical(self,"AssessmentBoy",f"{e}")
    def update_grade_subject(self):
        """Update grade and subject when combobox changes."""
        self.grade = self.gradecombo.currentText()
        self.subject = self.subjectcombo.currentText()

    def hidetotalscore(self):
        """Disable total score widgets if the subject's total already exists."""
        try:
            self.totalfield.clear()
            #checking whether total score has been filled before proceeding to
            #input scores
            cur.execute("""
                SELECT total_score FROM total 
                WHERE term_id = (SELECT term_id FROM term WHERE is_active = 1)
                AND grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND subject_id = (SELECT subject_id FROM subject WHERE subject_abbr = %s)
            """, (self.grade, self.subject))

            total = cur.fetchone()
            #checking where subject scores has already been filled
            cur.execute("""SELECT subject_score FROM score WHERE subject_id=(
                                                       SELECT subject_id FROM subject WHERE subject_abbr=%s)AND grade_id=
                                                       (SELECT grade_id FROM grade WHERE grade_name=%s)
                                                       AND term_id=(SELECT term_id FROM term
                                                       WHERE is_active=1)""",
                        (self.subject, self.grade))
            subject = cur.fetchall()

            if total:

                self.totalbutton.setDisabled(True)
                self.totalfield.setDisabled(True)
                self.totalfield.setText(str(total[0]))
                self.enterscoresbutton.setDisabled(False)


                if subject:
                    self.enterscoresbutton.setDisabled(True)
            else:
                self.totalbutton.setDisabled(False)
                self.totalfield.setDisabled(False)
                self.enterscoresbutton.setDisabled(True)
                if not subject:
                    self.enterscoresbutton.setDisabled(False)



        except Exception as e:

            QMessageBox.critical(self, "Error", "Failed to fetch total score.")

    def loadlearners(self, grade):
        """Loading learners detail from the database"""
        self.scorestable.setColumnWidth(0, 40)   # Column 0 (learner_id)
        self.scorestable.setColumnWidth(1, 250)  # Column 1 (name)
        self.scorestable.setColumnWidth(2, 40)  # Column 2 (score)

        cur.execute("""
            SELECT learner_id, first, second, surname
            FROM learner
            WHERE grade = %s
        """, (self.grade,))
        
        result = cur.fetchall()
        
        self.scorestable.setRowCount(len(result))
        
        for row, (learner_id, first, second, surname) in enumerate(result):
            name = f"{first} {second} {surname}"
            
            self.scorestable.setItem(row, 0, QTableWidgetItem(str(learner_id)))  # ID
            self.scorestable.setItem(row, 1, QTableWidgetItem(name))             # Full name
            self.scorestable.setItem(row, 2, QTableWidgetItem(""))               # Score (empty for now)

        self.scorestable.verticalHeader().setDefaultSectionSize(15)  # Or 25, 35 etc
        self.scorestable.verticalHeader().setFixedWidth(40)  # Adjust the number as needed
        self.hidetotalscore()

    def savetotalscore(self):
        """Saving total score per learning area/subject and per grade using upsert"""
        # subject = self.subjectcombo.currentText()
        try:
            total=self.totalfield.text()
            if len(total)==0:
                QMessageBox.information(self,"AssessmentBoy","Total Field is Required")
                return
            else:
        
                total = int(self.totalfield.text())

            # First get all the required IDs
            cur.execute("""
                SELECT 
                    (SELECT term_id FROM term WHERE is_active = 1 LIMIT 1) as term_id,
                    (SELECT subject_id FROM subject WHERE subject_abbr = %s LIMIT 1) as subject_id,
                    (SELECT grade_id FROM grade WHERE grade_name = %s LIMIT 1) as grade_id
            """, (self.subject, self.grade))

            ids = cur.fetchone()

            if ids and all(ids):  # Check all IDs exist
                term_id, subject_id, grade_id = ids

                # Use INSERT ON DUPLICATE KEY UPDATE
                cur.execute("""
                    INSERT INTO total (term_id, grade_id, subject_id, total_score)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE total_score = VALUES(total_score)
                """, (term_id, grade_id, subject_id, total))

                db.commit()
                QMessageBox.information(self,"AssessmentBoy","Total score saved successfully")
                self.totalfield.clear()
                self.hidetotalscore()

            else:
                QMessageBox.critical(self,"AssessmentBoy","Could not find all required IDs")

        except Exception as e:
            db.rollback()
            QMessageBox.critical(self,"AssessmentBoy",f"Error saving total score: {e}")
    def savescores(self):
        # Get IDs
        try: 
            cur.execute("SELECT grade_id FROM grade WHERE grade_name=%s", (self.grade,))
            grade_id = cur.fetchone()[0]
            
            cur.execute("SELECT subject_id FROM subject WHERE subject_abbr=%s", (self.subject,))
            subject_id = cur.fetchone()[0]
            
            cur.execute("SELECT term_id FROM term WHERE is_active=1")
            term_id = cur.fetchone()[0]

            # Validate total
            if not self.totalfield.text():
                QMessageBox.critical(self, "AssessmentBoy", "Please set the total score first")
                return
            total = int(self.totalfield.text())

            for row in range(self.scorestable.rowCount()):
                learner_id = int(self.scorestable.item(row, 0).text())
                score_text = self.scorestable.item(row, 2).text()
                
                if not score_text:
                    QMessageBox.critical(self, "AssessmentBoy", f"Please enter score for row {row+1}")
                    return
                    
                score = int(score_text)
                
                if score > total:
                    QMessageBox.critical(self, "AssessmentBoy", 
                                    f"Score {score} exceeds total {total} in row {row+1}")
                    return

                # Calculate percentage
                percentage_score = (score / total) * 100
                rounded_score = round(percentage_score)
                
                if percentage_score <= 29:
                    grading_level = "BE"
                elif percentage_score <= 49:
                    grading_level = "AE"
                elif percentage_score <= 99:
                    grading_level = "ME"
                else:
                    grading_level = "EE"

                try:
                    # Save/update subject score
                    cur.execute("""
                        INSERT INTO score(grade_id, learner_id, subject_id, term_id, 
                                        subject_score, expectation)
                        VALUES(%s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        subject_score = VALUES(subject_score),
                        expectation = VALUES(expectation)
                    """, (grade_id, learner_id, subject_id, term_id, 
                        rounded_score, grading_level))
                    
                    # Calculate FRESH grand total by summing all subjects for this term
                    cur.execute("""
                        SELECT COALESCE(SUM(subject_score), 0) 
                        FROM score 
                        WHERE learner_id = %s 
                        AND term_id = %s 
                        AND grade_id = %s
                    """, (learner_id, term_id, grade_id))
                    new_grandtotal = cur.fetchone()[0]
                    
                    # Update grand total
                    cur.execute("""
                        INSERT INTO grand(learner_id, term_id, grade_id, grandtotal)
                        VALUES(%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        grandtotal = VALUES(grandtotal)
                    """, (learner_id, term_id, grade_id, new_grandtotal))
                    
                    db.commit()
                    
                except Exception as e:
                    db.rollback()
                    QMessageBox.critical(self, "Database Error", 
                                    f"Failed to save scores: {str(e)}")
                    return

            # Clear fields
            for row in range(self.scorestable.rowCount()):
                self.scorestable.item(row, 2).setText("")
                
            self.enterstatusbar.showMessage("Scores saved successfully", 3000)
            self.hidetotalscore()
            self.displayentered()
            # dbfunc()
        except Exception as e:
            QMessageBox.critical(self,"AssessmentBoy",f"{e}")
        
    def homepage(self):
        home=mainwin()
        widget.setCurrentIndex(widget.currentIndex()+1)
        widget.addWidget(home)

class updatescore(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("updatescores.ui"), self)
        self.totalbutton.clicked.connect(self.loadlearnerlist)
        # self.grade = self.gradecombo.currentText()
        # self.subject = self.subjectcombo.currentText()
        self.gradecombo.currentTextChanged.connect(self.loadlearnerlist)
        self.listcombo.currentTextChanged.connect(self.loadlearnerscore)
        self.updatescorebutton.clicked.connect(self.updatescore)
        self.homebutton.clicked.connect(self.tohome)

    def loadlearnerlist(self):
        grade = self.gradecombo.currentText()
        subject = self.subjectcombo.currentText()
        self.listcombo.clear()

        cur.execute("""SELECT learner_id,first,second, surname FROM learner
        WHERE grade=%s""",(grade,))
        l=cur.fetchall()
        for i in l:
            learner=f"{i[0]}. {i[1]} {i[2]}"

            self.listcombo.addItem(str(learner))

    def loadlearnerscore(self):
        self.scorefield.clear()
        #retrieving total score
        cur.execute("""SELECT total_score FROM total WHERE term_id=(SELECT term_id FROM
        term WHERE is_active=1) AND subject_id=(SELECT subject_id FROM subject WHERE subject_abbr=%s)
        AND grade_id=(SELECT grade_id FROM grade WHERE grade_name=%s)""",(self.subjectcombo.currentText(),
                                                                          self.gradecombo.currentText()))
        total=cur.fetchone()
        if total:
            total_score=total[0]
            total_score=int(total_score)
        selected = self.listcombo.currentText()
        learner_id = selected.split(".")[0].strip()
        # learner_id=int(learner_id)
        subject_abbr=self.subjectcombo.currentText()
        cur.execute("""SELECT subject_score FROM score WHERE subject_id=(
        SELECT subject_id FROM subject WHERE subject_abbr=%s
        AND learner_id=%s)""",(subject_abbr,learner_id))
       
        s=cur.fetchone()
        if s:
            score=s[0]
            raw_score=int(score)/100*int(total_score)
            raw_score=round(raw_score)
            self.scorefield.setText(str(raw_score))
            self.scorefield.setDisabled(False)
            self.updatescorebutton.setDisabled(False)
        else:
            self.scorefield.setDisabled(True)
            self.updatescorebutton.setDisabled(True)

    def updatescore(self):
        try:
            # Get learner ID from combo box
            learner_id = int(self.listcombo.currentText().split(".")[0])
            
            # Validate score input
            score = self.scorefield.text()  # Changed from undefined 'score' variable
            if not score:
                QMessageBox.information(self, "AssessmentBoy", "Score field is required")
                return
            
            score = int(score)
            
            # Get subject ID
            cur.execute("SELECT subject_id FROM subject WHERE subject_abbr=%s", 
                    (self.subjectcombo.currentText(),))
            subject_id = cur.fetchone()[0]
            
            # Get total score for comparison
            cur.execute("""
                SELECT total_score FROM total 
                WHERE term_id = (SELECT term_id FROM term WHERE is_active=1) 
                AND subject_id = %s
                AND grade_id = (SELECT grade_id FROM grade WHERE grade_name=%s)
            """, (subject_id, self.gradecombo.currentText()))
            
            total_result = cur.fetchone()
            if not total_result:
                QMessageBox.critical(self, "AssessmentBoy", "Total score not found for this subject")
                return
                
            total_score = int(total_result[0])
            
            # Calculate percentage score
            subject_score = round((score / total_score) * 100)
            if subject_score > 100:
                QMessageBox.critical(self, "AssessmentBoy", "Score cannot exceed total score")
                return
            
            # Update subject score
            cur.execute("""
                UPDATE score 
                SET subject_score = %s 
                WHERE learner_id = %s 
                AND subject_id = %s
                AND term_id = (SELECT term_id FROM term WHERE is_active=1)
            """, (subject_score, learner_id, subject_id))
            
            # Update grand total (properly calculated)
            cur.execute("""
                SELECT COALESCE(SUM(subject_score), 0) 
                FROM score 
                WHERE learner_id = %s
                AND term_id = (SELECT term_id FROM term WHERE is_active=1)
            """, (learner_id,))
            
            newgrandtotal = cur.fetchone()[0]
            
            cur.execute("""
                UPDATE grand 
                SET grandtotal = %s 
                WHERE learner_id = %s
                AND term_id = (SELECT term_id FROM term WHERE is_active=1)
            """, (newgrandtotal, learner_id))
            
            db.commit()
            self.scorefield.clear()
            self.updatestatusbar.showMessage("Score updated successfully", 3000)
            
        except ValueError:
            QMessageBox.critical(self, "AssessmentBoy", "Please enter valid numeric values")
        except Exception as e:
            db.rollback()
            QMessageBox.critical(self, "AssessmentBoy", f"Error updating score: {str(e)}")
    
    def tohome(self):
        screen=mainwin()
        widget.addWidget(screen)
        widget.addWidget(screen)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
class deletescore(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("deletescores.ui"), self)
        
        self.subjectdelete.clicked.connect(self.deletesubject)
        self.homebutton.clicked.connect(self.tohome)
        self.gradedelete.clicked.connect(self.deletegradescore)
        self.subdelete.clicked.connect(self.deletelsubject)
        self.deleteall.clicked.connect(self.deletelall)
        
    def deletesubject(self):
        resp = QMessageBox.question(
            self,
            "AssessmentBoy",
            f"Are you sure you want to delete\n{self.subjectcombo.currentText()} for Grade {self.gradecombo.currentText()}?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if resp != QMessageBox.Yes:
            return  # Exit if user clicks "No"

        try:
            grade_name = self.gradecombo.currentText()
            subject_abbr = self.subjectcombo.currentText()

            # 1. FIRST, fetch the subject_score BEFORE deleting it
            cur.execute("""
                SELECT subject_score FROM score
                WHERE subject_id = (SELECT subject_id FROM subject WHERE subject_abbr = %s)
                AND grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (subject_abbr, grade_name))

            subject_score_result = cur.fetchone()
            if subject_score_result:  # Only proceed if the score exists
                subject_score = subject_score_result[0]

                # 2. Fetch the current grand total
                cur.execute("""
                    SELECT grandtotal FROM grand 
                    WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s) 
                    AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
                """, (grade_name,))
                
                grandtotal_result = cur.fetchone()
                if grandtotal_result:
                    grandtotal = grandtotal_result[0]
                    new_grand = grandtotal - subject_score
                    # 3. Update the grand total BEFORE deleting the score
                    cur.execute("""
                        UPDATE grand 
                        SET grandtotal = %s
                        WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                        AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
                    """, (new_grand, grade_name))

            # 4. NOW delete the score and related records
            cur.execute("""
                DELETE FROM score 
                WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND subject_id = (SELECT subject_id FROM subject WHERE subject_abbr = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (grade_name, subject_abbr))

            cur.execute("""
                DELETE FROM total 
                WHERE term_id = (SELECT term_id FROM term WHERE is_active = 1)
                AND grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND subject_id = (SELECT subject_id FROM subject WHERE subject_abbr = %s)
            """, (grade_name, subject_abbr))

            db.commit()  
            self.deletestatusbar.showMessage("Scores deleted successfully", 3000)

        except Exception as e:
            db.rollback()  # Revert changes on error
            QMessageBox.critical(self,"AssessmentBoy",f"Database error: {e}")
    
    def deletegradescore(self):
        # Ask for confirmation
        confirm = QMessageBox.question(
            self,
            "AssessmentBoy",  # Title
            f"Are you sure you want to delete all\nscores for Grade {self.gradecombo.currentText()}?",  # Message
            QMessageBox.Yes | QMessageBox.No,  # Buttons
            QMessageBox.No  # Default button (avoids accidental deletion)
        )

        if confirm != QMessageBox.Yes:
            return  # Exit if user clicks "No"

        try:
            grade_name = self.gradecombo.currentText()

            # 1. Delete from score table
            cur.execute("""
                DELETE FROM score 
                WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (grade_name,))

            # 2. Delete from grand table
            cur.execute("""
                DELETE FROM grand 
                WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (grade_name,))

            # 3. Delete from total table
            cur.execute("""
                DELETE FROM total 
                WHERE grade_id = (SELECT grade_id FROM grade WHERE grade_name = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (grade_name,))

            db.commit()
            QMessageBox.information(
                self,
                "AssessmentBoy",
                f"Scores for Grade {grade_name}\ndeleted successfully!",
                QMessageBox.Ok
            )

        except Exception as e:
            db.rollback()
            QMessageBox.critical(
                self,
                "AssessmentBoy - Error",
                f"Failed to delete scores:\n{str(e)}",
                QMessageBox.Ok
            )
    def deletelsubject(self):
        try:
            learner_id = int(self.delfield.text())
            
            # Get learner details
            cur.execute("SELECT first, second, surname FROM learner WHERE learner_id=%s", (learner_id,))
            name = cur.fetchone()
            if not name:
                QMessageBox.warning(self, "AssessmentBoy", "Learner not found")
                return
                
            full_name = f"{name[0]} {name[1]} {name[2]}"
            
            # Confirmation dialog
            resp = QMessageBox.question(
                self,
                "AssessmentBoy",
                f"Are you sure you want to delete\n{self.subjectcombo.currentText()} for {full_name}?",
                QMessageBox.Yes | QMessageBox.No
            )
            if resp != QMessageBox.Yes:
                return

            # SOFT DELETE: Update score to 0 instead of deleting
            cur.execute("""
                UPDATE score 
                SET subject_score = 0 
                WHERE learner_id = %s
                AND subject_id = (SELECT subject_id FROM subject WHERE subject_abbr = %s)
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (learner_id, self.subjectcombo.currentText()))

            # Update grand total (sum will now exclude this subject)
            cur.execute("""
                SELECT COALESCE(SUM(subject_score), 0) 
                FROM score 
                WHERE learner_id = %s 
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (learner_id,))
            new_grandtotal = cur.fetchone()[0]

            cur.execute("""
                UPDATE grand 
                SET grandtotal = %s
                WHERE learner_id = %s 
                AND term_id = (SELECT term_id FROM term WHERE is_active = 1)
            """, (new_grandtotal, learner_id))

            db.commit()
            QMessageBox.information(self, "AssessmentBoy", "Subject score delete successfully!")
            self.delfield.clear()
        except ValueError:
            QMessageBox.critical(self, "AssessmentBoy", "Invalid learner ID")
        except Exception as e:
            db.rollback()
            QMessageBox.critical(self, "AssessmentBoy", f"Error: {str(e)}")
            
#deleting all learning areas for a learner
    def deletelall(self):
        try:
            # Get learner ID from input field
            learner_id = int(self.delfield.text())
            
            # Access learner details
            cur.execute("""
                SELECT first, second, surname FROM learner
                WHERE learner_id = %s
            """, (learner_id,))
            name = cur.fetchone()
            
            if not name:
                QMessageBox.warning(self, "AssessmentBoy", "Learner not found")
                return
                
            full_name = f"{name[0]} {name[1]} {name[2]}"
            
            # Confirmation dialog
            resp = QMessageBox.question(
                self,
                "AssessmentBoy",
                f"Are you sure you want to CLEAR ALL\nSCORES for {full_name}?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No  # Default to 'No' for safety
            )
            
            if resp != QMessageBox.Yes:
                return

            # Get active term
            cur.execute("SELECT term_id FROM term WHERE is_active = 1")
            term_id = cur.fetchone()[0]

            # Soft delete all subject scores (set to 0)
            cur.execute("""
                UPDATE score 
                SET subject_score = 0 
                WHERE learner_id = %s
                AND term_id = %s
            """, (learner_id, term_id))

            # Update grand total to 0
            cur.execute("""
                UPDATE grand 
                SET grandtotal = 0
                WHERE learner_id = %s 
                AND term_id = %s
            """, (learner_id, term_id))

            db.commit()
            
            # Refresh UI if needed
            self.delfield.clear()
            QMessageBox.information(
                self, 
                "AssessmentBoy", 
                f"All scores for {full_name} cleared\nsuccessfully!",
                QMessageBox.Ok
            )
            
        except ValueError:
            QMessageBox.critical(self, "AssessmentBoy", "Please enter a valid learner ID")
        except Exception as e:
            db.rollback()
            QMessageBox.critical(
                self, 
                "AssessmentBoy", 
                f"Error clearing scores:\n{str(e)}"
            )
        
    def tohome(self):
        screen=mainwin()
        widget.addWidget(screen)
        widget.addWidget(screen)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
class adminpanel(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("admin.ui"), self)
        self.homebutton.clicked.connect(self.tohome)
        # self.testcombo.currentTextChanged.connect(self.saveassessment)
        self.adminchange.clicked.connect(self.saveassessment)
        self.datesave.clicked.connect(self.openclosedate)
        self.saveschool.clicked.connect(self.selectschool)
        self.loadterm()
        
        
    def selectschool(self):
        school_path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
        school=self.setschoolcombo.currentText()
        with open(school_path,"w") as file:
            file.write(school)
        self.adminstatusbar.showMessage(f"{school} saved successfully!",3000)
    def saveassessment(self):
        test=self.testcombo.currentText()
        path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/term.text"
        #selected term
        with open(path,"w")as file:
            file.write(test)
        cur.execute("""UPDATE term SET is_active=0
                    WHERE is_active=1""")
        cur.execute("""UPDATE term SET is_active=1 
                    WHERE selected_term=%s AND
                    selected_year=%s""",
                    (self.termcombo.currentText(),
                    self.yearcombo.currentText()))
        db.commit()
        # cur.execute("SELECT term_id FROM term WHERE is_active=1")
        # for items in cur.fetchall():
        #     print(items[0])
        self.adminstatusbar.showMessage("Changes Saved Successfully",3000)
        self.loadterm()
        
    def tohome(self):
        screen=mainwin()
        widget.addWidget(screen)
        widget.addWidget(screen)
        widget.setCurrentIndex(widget.currentIndex()+1)
    
    def openclosedate(self):
        #setting closing date
        closedate=self.closedate.date()
        closing_date=closedate.toString("dd-MMM-yyyy")
        path="D:/TONNIEGIFTED/Documents/programs/Remedial2/closingdate.txt"
        with open(path,"w") as file:
            file.write(closing_date)
        #setting closing date
        opendate=self.closedate.date()
        opening_date=opendate.toString("dd-MMM-yyyy")
        path="D:/TONNIEGIFTED/Documents/programs/Remedial2/openingdate.txt"
        with open(path,"w") as file:
            file.write(opening_date)
        
        self.adminstatusbar.showMessage("Dates saved successfully",3000)
    
    def loadterm(self):
            cur.execute("""SELECT selected_term,selected_year
                        FROM term WHERE is_active=1""")
            term=cur.fetchone()
            term_year=f"Term {term[0]}, {term[1]}"
            self.termlabel.setText(term_year) 
            #loading test
            path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/term.text"
            with open(path,"r")as file:
                test=file.read()
            self.testlabel.setText(f"{test} Assessment")
            
#adding learners
class savelearner(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("enterlearner.ui"), self)
        self.homebutton.clicked.connect(self.tohome)
        self.savebutton.clicked.connect(self.savelearner)
        self.gradecombo.currentTextChanged.connect(self.loadlearners)
        self.loadlearners()
        self.updatebutton.clicked.connect(self.updatelearner)
        self.deletebutton.clicked.connect(self.deletelearner)
        self.setTabOrder(self.firstfield, self.secondfield)
        self.setTabOrder(self.secondfield, self.surnamefield)
        
    def loadlearners(self):
        self.learnercombo.clear()  # Clear existing items first
        
        cur.execute("""SELECT learner_id, first, second, surname
                    FROM learner WHERE grade=%s""",
                    (self.gradecombo.currentText(),))
        
        learners = cur.fetchall()
        
        for learner_id, first, second, surname in learners:
            learner = f"{learner_id}. {first} {second}"
            if surname:  # Only add surname if it exists
                learner += f" {surname}"
            self.learnercombo.addItem(learner)
            
    def updatelearner(self):
        first=self.firstfield.text().strip().title()
        second=self.secondfield.text().strip().title()
        surname=self.surnamefield.text().strip().title()
        grade= self.gradecombo.currentText()
        learner_id=int(self.learnercombo.currentText().split(".")[0])
        learner=self.learnercombo.currentText().split(".")[1]
        # learner=(f"{first} {second} {surname}")
        if all([first,second,grade,learner_id]):
            reply = QMessageBox.question(
                    self,
                    "AssessmentBoy",
                    f"Do you want to update\n{learner} details?",
                    QMessageBox.Yes | QMessageBox.No
                )
            if reply==QMessageBox.Yes:
                cur.execute("""UPDATE learner SET first=%s,second=%s
                            ,surname=%s,grade=%s WHERE learner_id=%s""",
                            (first,second,surname,grade,learner_id))
                db.commit()
                self.learnerstatus.showMessage("Learner's details updated successfully!")
                self.loadlearners()
                self.idfield.clear()
                self.firstfield.clear()
                self.secondfield.clear()
                self.surnamefield.clear()
                self.loadlearners()
        else:
            QMessageBox.warning(self,"AssessmentBoy","Input all the necessary fields")
            return
        
    def savelearner(self):
        # Get all field values (strip whitespace)
        id_text = self.idfield.text().strip()
        first = self.firstfield.text().strip().title()
        second = self.secondfield.text().strip().title()
        surname = self.surnamefield.text().strip().title()  # Optional (can be empty)
        grade = self.gradecombo.currentText()

        # Validate REQUIRED fields (ID, first, second, grade)
        if not all([id_text, first, second, grade]):  # surname not checked here
            QMessageBox.critical(self,"AssessmentBoy"," ID, First, Second, and Grade are required")
            return

        try:
            learner_id = int(id_text)  # Convert ID to integer
        except ValueError:
            self.learnerstatus.showMessage("Error: ID must be a number", 3000)
            return

        try:
            # Check if learner exists
            cur.execute("SELECT first, second, surname FROM learner WHERE learner_id = %s", (learner_id,))
            existing = cur.fetchone()

            if existing:  # Learner exists → UPDATE after confirmation
                existing_name = f"{existing[0]} {existing[1]} {existing[2] if existing[2] else ''}".strip()
                reply = QMessageBox.question(
                    self,
                    "AssessmentBoy",
                    f"ID {learner_id} already exists do you want\nto overwrite{existing_name}\ndetails?",
                    QMessageBox.Yes | QMessageBox.No
                )

                if reply == QMessageBox.Yes:
                    cur.execute("""
                        UPDATE learner 
                        SET first = %s, second = %s, surname = %s, grade = %s
                        WHERE learner_id = %s
                    """, (first, second, surname if surname else None, grade, learner_id))  # Handle NULL surname
                    self.learnerstatus.showMessage("Learner updated successfully!", 3000)
                    self.idfield.clear()
                    self.firstfield.clear()
                    self.secondfield.clear()
                    self.surnamefield.clear()
                    self.loadlearners()
                    
            else:  # New learner → INSERT
                cur.execute("""
                    INSERT INTO learner 
                    (learner_id, first, second, surname, grade)
                    VALUES (%s, %s, %s, %s, %s)
                """, (learner_id, first, second, surname if surname else None, grade))  # Handle NULL surname
                self.learnerstatus.showMessage("New learner added successfully!", 3000)
                self.idfield.clear()
                self.firstfield.clear()
                self.secondfield.clear()
                self.surnamefield.clear()
                self.loadlearners()
            db.commit()

        except Exception as e:
            db.rollback()
            self.learnerstatus.showMessage(f"Database error: {str(e)}", 3000)
            
    def deletelearner(self):
        learner_id=int(self.learnercombo.currentText().split(".")[0])
        learner=self.learnercombo.currentText().split(".")[1]
        reply = QMessageBox.question(
                self,
                "AssessmentBoy",
                f"Do you want to update\n{learner} details?",
                QMessageBox.Yes | QMessageBox.No
            )
        if reply==QMessageBox.Yes:
            cur.execute("SET FOREIGN_KEY_CHECKS=0")
            cur.execute("""DELETE FROM learner WHERE learner_id=%s""",(learner_id,))
            self.learnerstatus.showMessage("Learner Delete successfully!",3000)
            self.loadlearners()
            cur.execute("SET FOREIGN_KEY_CHECKS=1")
            db.commit()
         
    def tohome(self):
        screen=mainwin()
        widget.addWidget(screen)
        widget.addWidget(screen)
        widget.setCurrentIndex(widget.currentIndex()+1)
class analysis(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi(resource_path("grading.ui"), self)
        self.homebutton.clicked.connect(self.tohome)
        self.analysebutton.clicked.connect(self.generate_assessment_report)  # Connect analyze button
        self.generatebutton.clicked.connect(self.on_generate_reports_clicked)
        self.loadtest()
        self.loadterm()
        self.readschool()
        
        # Database configuration
        self.DB_CONFIG = {
            'host': 'localhost',
            'user': 'root',
            'password': 'print',
            'database': 'assess'
        }

     #display grades     
    def readschool(self):
        school_path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
        with open(school_path,"r") as file:
            school=file.read()
            # self.level_label.setText(school)
        if school =="Junior School":
            grades=['Seven','Eight','Nine']
        elif school=="Upper Primary":
            grades=['Four','Five','Six']
        else:
            grades=['One','Two','Three']   
            
        self.gradecombo.addItems(grades)
        self.schoollabel.setText(school)
    def tohome(self):
        screen = mainwin()
        widget.addWidget(screen)
        widget.setCurrentIndex(widget.currentIndex() + 1)

    def calculate_deviations(self):
        """Calculate term-to-term grandtotal deviations."""
        try:
            db = mysql.connector.connect(**self.DB_CONFIG)
            cur = db.cursor()

            # Current term (term_id=20)
            cur.execute("""
                SELECT learner_id, grandtotal FROM grand 
                WHERE grade_id=(SELECT grade_id FROM grade
                WHERE grade_name=%s) AND term_id=(
                SELECT term_id FROM term WHERE is_active=
                1
                )
            """,(self.gradecombo.currentText(),))
            current = {row[0]: row[1] for row in cur.fetchall()}

            # Previous term (term_id=19)
            cur.execute("""
                SELECT learner_id, grandtotal FROM grand 
                WHERE grade_id=(SELECT grade_id FROM grade
                WHERE grade_name=%s) AND term_id=(SELECT
                term_id FROM term WHERE term_number=(SELECT
                term_number-1 FROM term WHERE is_active=1))
            """, (self.gradecombo.currentText(),))  # Added missing parameter
            previous = {row[0]: row[1] for row in cur.fetchall()}

            return {lid: current[lid] - previous.get(lid, 0) for lid in current}

        except Exception as e:
            QMessageBox.critical(self,"AssessmentBoy",f"Deviation calculation error: {e}")
            return {}
        finally:
            if 'db' in locals() and db.is_connected():
                cur.close()
                db.close()
    def calculate_grade(self, score):
        """Convert score to grade (BE/AE/ME/EE).
        BE: Below Expectations (≤261)
        AE: Approaching Expectations (262-441)
        ME: Meeting Expectations (442-891)
        EE: Exceeding Expectations (≥892)
        """
        try:
            score = float(score)
            if score <= 261: return "BE"
            elif score <= 441: return "AE"
            elif score <= 891: return "ME"
            return "EE"
        except (TypeError, ValueError):
            return "INVALID"  # Or raise an exception

    def fetch_learner_data(self):
        """Fetch all learner data from database."""
        try:
            db = mysql.connector.connect(**self.DB_CONFIG)
            cur = db.cursor()

            # Get ranked learners with full names and admission numbers
            cur.execute("""
                SELECT l.learner_id, l.learner_id, 
                    CONCAT(COALESCE(l.first, ''), ' ', 
                            COALESCE(l.second, ''), ' ', 
                            COALESCE(l.surname, '')) AS fullname, 
                    g.grandtotal
                FROM learner l JOIN grand g ON l.learner_id=g.learner_id
                WHERE g.grade_id=(SELECT grade_id FROM grade WHERE 
                grade_name=%s) 
                AND g.term_id=(SELECT term_id FROM term
                WHERE is_active=1)
                ORDER BY g.grandtotal DESC
            """,(self.gradecombo.currentText(),))
            learners = [
                {
                    "id": row[0],
                    "adm": row[1],  # Admission number
                    "name": row[2], 
                    "gt": row[3], 
                    "pos": i+1,  # Position based on ranking
                    "gt_ex": self.calculate_grade(row[3])
                }
                for i, row in enumerate(cur.fetchall())
            ]

            # Get subject scores
            cur.execute("""
                SELECT s.learner_id, b.subject_abbr, s.subject_score, s.expectation
                FROM score s JOIN subject b ON s.subject_id=b.subject_id
                WHERE s.grade_id=(SELECT grade_id FROM grade WHERE grade_name=%s) 
                AND s.term_id=(SELECT term_id FROM term WHERE is_active=1)
            """, (self.gradecombo.currentText(),))
            subjects = {}
            for row in cur.fetchall():
                if row[0] not in subjects:
                    subjects[row[0]] = {}
                subjects[row[0]][row[1]] = (row[2], row[3])

            return learners, subjects

        except Exception as e:
            QMessageBox.critical(self,"AssessmentBoy",f"Database error: {e}")
            return [], {}
        finally:
            if 'db' in locals() and db.is_connected():
                cur.close()
                db.close()

    def create_cell(self, ws, row, col, value, font_size=12, bold=False, merge=False, fill=None, center=False):
        """Create a cell with left alignment (or centered if specified) and optional merging."""
        if merge:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+merge-1)
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border
        return cell

    def generate_assessment_report(self):
        """Generate assessment report with direct cell styling (no NamedStyle)."""
        try:
            # 1. Get school level
            with open("D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text", "r") as f:
                school = f.read().strip()

            # 2. Create workbook with print settings
            wb = Workbook()
            ws = wb.active
            ws.title = "Assessment Analysis"
            
            # Configure print settings
            ws.print_options.gridLines = True
            ws.sheet_view.showGridLines = True
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=1.0)
            ws.print_options.horizontalCentered = True

            # 3. Write report headers with direct styling
            ws.merge_cells('A1:Y1')
            ws['A1'] = "IGAMBA JUNIOR AND PRIMARY SCHOOL"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            grade_info = f"GRADE {self.gradecombo.currentText()} {self.testlabel.text()}, {self.termlabel.text()}"
            ws.merge_cells('A2:Y2')
            ws['A2'] = grade_info
            ws['A2'].font = Font(bold=True, size=11)
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A3:Y3')
            ws['A3'] = "ASSESSMENT ANALYSIS"
            ws['A3'].font = Font(bold=True, size=11)
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

            # 4. Get and validate data
            learners, subjects = self.fetch_learner_data()
            if not learners:
                QMessageBox.warning(self, "Warning", "No learner data found in database")
                return

            deviations = self.calculate_deviations()

            # 5. Configure columns based on school level
            if school == 'Lower Primary':
                headers = ["ADM", "FULLNAME", "POS"] + \
                        sum([[subj, "EX"] for subj in ["EN/ACT", "MA/ACT", "KI/ACT", "RE/ACT", "ENV/ACT", "C/ACT"]], []) + \
                        ["GT", "EX", "DEV"]
                subject_cols = ["EN/ACT", "MA/ACT", "KI/ACT", "RE/ACT", "ENV/ACT", "C/ACT"]
                widths = {'A':5, 'B':25, 'C':5, 'D':5, 'E':4, 'F':5, 'G':4, 'H':5, 'I':4,
                        'J':5, 'K':4, 'L':5, 'M':4, 'N':5, 'O':4, 'P':6, 'Q':4, 'R':6}
            elif school == 'Upper Primary':
                headers = ["ADM", "FULLNAME", "POS"] + \
                        sum([[subj, "EX"] for subj in ["ENG", "MATHS", "KISW", "SST", "AGN", "SCIE", "C/A", "CRE"]], []) + \
                        ["GT", "EX", "DEV"]
                subject_cols = ["ENG", "MATHS", "KISW", "SST", "AGN", "SCIE", "C/A", "CRE"]
                widths = {'A':5, 'B':25, 'C':5, 'D':5, 'E':4, 'F':6, 'G':4, 'H':5, 'I':4,
                        'J':5, 'K':4, 'L':5, 'M':4, 'N':5, 'O':4, 'P':5, 'Q':4, 'R':5, 'S':4,
                        'T':6, 'U':4, 'V':6}
            else:  # Junior School
                headers = ["ADM", "FULLNAME", "POS"] + \
                        sum([[subj, "EX"] for subj in ["ENG", "MATHS", "KISW", "INT", "SST", "AGN", "PTC", "CAS", "CRE"]], []) + \
                        ["GT", "EX", "DEV"]
                subject_cols = ["ENG", "MATHS", "KISW", "INT", "SST", "AGN", "PTC", "CAS", "CRE"]
                widths = {'A':5, 'B':25, 'C':5, 'D':5, 'E':4, 'F':6, 'G':4, 'H':5, 'I':4,
                        'J':5, 'K':4, 'L':5, 'M':4, 'N':5, 'O':4, 'P':5, 'Q':4, 'R':5, 'S':4,
                        'T':5, 'U':4, 'V':6, 'W':4, 'X':6}

            # 6. Write column headers with direct styling
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(bottom=Side(style='medium'))

            # 7. Write student data with direct styling
            data_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_num, learner in enumerate(learners, 6):
                # Basic info
                for col, val in [(1, learner["adm"]), (2, learner["name"]), (3, learner["pos"])]:
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.border = data_border
                
                # Subject scores
                col = 4
                for subject in subject_cols:
                    score, ex = subjects.get(learner["id"], {}).get(subject, ("", ""))
                    ws.cell(row=row_num, column=col, value=score).border = data_border
                    ws.cell(row=row_num, column=col+1, value=ex).border = data_border
                    col += 2
                
                # Totals
                ws.cell(row=row_num, column=col, value=learner["gt"]).border = data_border
                ws.cell(row=row_num, column=col+1, value=learner["gt_ex"]).border = data_border
                ws.cell(row=row_num, column=col+2, value=deviations.get(learner["id"], 0)).border = data_border

            # 8. Set column widths
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            # 9. Add timestamp footer
            footer_row = ws.max_row + 2
            timestamp = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
            footer_cell = ws.cell(row=footer_row, column=1, value=timestamp)
            footer_cell.alignment = Alignment(horizontal='center')
            footer_cell.font = Font(size=9, italic=True)

            # 10. Save with file dialog
            default_name = f"Assessment_{self.gradecombo.currentText()} {self.termlabel.text()}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Assessment Report",
                f"D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/{default_name}",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                    
                if os.path.exists(file_path):
                    reply = QMessageBox.question(
                        self,
                        "Overwrite File?",
                        f"'{os.path.basename(file_path)}' already exists. Overwrite?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply != QMessageBox.Yes:
                        return

                try:
                    wb.save(file_path)
                    QMessageBox.information(
                        self,
                        "Success",
                        f"Analysis saved successfully!"
                    )
                except PermissionError:
                    QMessageBox.critical(
                        self,
                        "Error",
                        "Permission denied. Please close the file if it's open elsewhere."
                    )
                except Exception as e:
                    QMessageBox.critical(
                        self,
                        "Error",
                        f"Failed to save file: {str(e)}"
                    )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An unexpected error occurred:\n{str(e)}")
    def loadtest(self):
        path="D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/term.text"
        with open(path,"r")as file:
            test=file.read()
        self.testlabel.setText(f"{test} Assessment")
        
    def loadterm(self):
            cur.execute("""SELECT selected_term,selected_year
                        FROM term WHERE is_active=1""")
            term=cur.fetchone()
            term_year=f"Term {term[0]}, {term[1]}"
            self.termlabel.setText(term_year) 
            # self.termlabel.setStyleSheet('color:blue',)
            
    #generating report sheets
    #  self.generate_button.clicked.connect(self.on_generate_reports_clicked)

    def on_generate_reports_clicked(self):
        """This method will be called when the button is clicked"""
        # Prepare parameters
        school_path = "D:/TONNIEGIFTED/Documents/programs/Remedial2/name.txt"
        path = "D:/TONNIEGIFTED/Documents/programs/Remedial2/closingdate.txt"
        
        try:
            # Read school info
            with open(path, "r") as file:
                closingdate = file.read()
            with open("D:/TONNIEGIFTED/Documents/programs/Remedial2/openingdate.txt", "r") as file:
                openingdate = file.read()
            with open(school_path, "r") as file:
                school_name = file.read()
                
            school_info = {
                'name': f'{school_name}'.upper(),
                'address': 'P.O. BOX 32-01003 GITUAMBA',
                'email': 'igambacomprehensive@gmail.com',
                'closing_date': closingdate,
                'opening_date': openingdate
            }
            
            logo_path = "D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/LOGO FINALE.png"
            
            # Get grade ID
            cur.execute("SELECT grade_id FROM grade WHERE grade_name=%s",
                    (self.gradecombo.currentText(),))
            grade_id = cur.fetchone()[0]
            
            # Generate reports
            generate_report_books(
                grade_id=grade_id,
                logo_path=logo_path,
                school_info=school_info
            )
            
            # QMessageBox.information(
            #     self, 
            #     "Success", 
            #     "Report sheets generated successfully"
            # )
            
        except FileNotFoundError as e:
            QMessageBox.critical(
                self,
                "File Error",
                f"Could not read required file: {str(e)}"
            )
        except mysql.connector.Error as db_error:
            QMessageBox.critical(
                self, 
                "Database Error", 
                f"Database operation failed: {str(db_error)}"
            )
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Error", 
                f"Failed to generate reports: {str(e)}"
            )
def generate_report_books(grade_id, logo_path, school_info):
    """
    Generate professional one-page-per-learner PDF report in a single document.
    Includes proper error handling and validation for all cases.
    """
    try:
        # 1. Validate school_info dictionary
        if not isinstance(school_info, dict) or not all(key in school_info for key in ['name', 'address', 'email', 'closing_date', 'opening_date']):
            raise ValueError("Invalid school_info dictionary provided")

        # 2. Verify logo file exists
        if not os.path.exists(logo_path):
            raise FileNotFoundError(f"Logo file not found at: {logo_path}")

        # 3. Get school level from file with validation
        school_path = "D:/TONNIEGIFTED/Documents/programs/AssessmentBoy/school.text"
        if not os.path.exists(school_path):
            raise FileNotFoundError(f"School level file not found at: {school_path}")
        
        with open(school_path, "r") as file:
            school = file.read().strip()
        
        valid_schools = ['Lower Primary', 'Upper Primary', 'Junior School']
        if school not in valid_schools:
            school = 'Junior School'  # Default if invalid value
            QMessageBox.warning(None, "Warning", f"Invalid school level detected. Defaulting to 'Junior School'")

        # 4. Set school-specific values
        if school == 'Lower Primary':
            grand_total_out_of = 600
            subject_out_of = 100
        elif school == 'Upper Primary':
            grand_total_out_of = 800
            subject_out_of = 100
        else:  # Junior School
            grand_total_out_of = 900
            subject_out_of = 100

        # 5. Database connection with error handling
        try:
            db = mysql.connector.connect(
                host="localhost",
                user="root",
                password="print",
                database="assess"
            )
            cur = db.cursor()
        except Exception as db_error:
            raise ConnectionError(f"Database connection failed: {str(db_error)}")

        # 6. Get active term with validation
        cur.execute("SELECT term_id, selected_term, selected_year FROM term WHERE is_active=1")
        term_data = cur.fetchone()
        
        if not term_data:
            db.close()
            raise ValueError("No active term found. Please set an active term first.")
        
        term_id, selected_term, selected_year = term_data
        term_year = f"Term {selected_term}, {selected_year}"

        # 7. Get all learners with their grand totals
        cur.execute("""SELECT l.learner_id, CONCAT(l.first,' ',l.second,' ',l.surname), g.grandtotal
                    FROM learner l JOIN grand g ON l.learner_id=g.learner_id
                    WHERE g.grade_id=%s AND g.term_id=%s
                    ORDER BY g.grandtotal DESC""", 
                    (grade_id, term_id))
        
        learners_data = cur.fetchall()
        
        if not learners_data:
            db.close()
            QMessageBox.warning(None, "No Data", "No learners found for the selected grade and term.")
            return  # Exit gracefully if no learners found

        class_size = len(learners_data)

        # 8. Get grade name with validation
        cur.execute("SELECT grade_name FROM grade WHERE grade_id=%s", (grade_id,))
        grade_result = cur.fetchone()
        
        if not grade_result:
            db.close()
            raise ValueError(f"No grade found with ID: {grade_id}")
        
        grade_name = grade_result[0]

        # 9. Create PDF document
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        
        for position, (learner_id, learner_name, grandtotal) in enumerate(learners_data, start=1):
            # Validate learner data
            if not learner_name or not isinstance(learner_name, str):
                learner_name = "Name Not Available"
            
            try:
                pdf.add_page()
                pdf.set_margins(20, 10, 20)
                
                # Header Section with validation
                try:
                    pdf.image(logo_path, x=20, y=10, w=30)
                except:
                    pass  # Skip logo if there's an error but continue with report
                
                pdf.set_font("Arial", 'B', 14)
                pdf.cell(0, 10, str(school_info.get('name', '')), 0, 1, 'C')
                pdf.set_font("Arial", size=10)
                pdf.cell(0, 5, str(school_info.get('address', '')), 0, 1, 'C')
                pdf.cell(0, 5, f"Email: {str(school_info.get('email', ''))}", 0, 1, 'C')
                pdf.ln(8)
                
                # Learner Info Section
                col1_width, col2_width = 30, 60
                pdf.ln(3)
                
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "ADM NO:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(col2_width, 8, str(learner_id), 0, 0)
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "NAME:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(0, 8, learner_name.upper(), 0, 1)
                
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "POSITION:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(col2_width, 8, str(position), 0, 0)
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "OUT OF:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(0, 8, str(class_size), 0, 1)
                
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "GRADE:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(col2_width, 8, str(grade_name), 0, 0)
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(col1_width, 8, "TERM:", 0, 0)
                pdf.set_font("Arial", '', 12)
                pdf.cell(0, 8, term_year, 0, 1)
                pdf.ln(8)
                
                # Subjects Table
                pdf.set_font("Arial", 'B', 10)
                col_widths = [70, 20, 20, 60]
                headers = ["LEARNING AREA", "SCORE", "OUT OF", "COMMENT"]
                
                for width, header in zip(col_widths, headers):
                    pdf.cell(width, 8, header, 1, 0, 'C')
                pdf.ln()
                
                # Get subject scores with error handling
                try:
                    cur.execute("""SELECT b.subject_id, b.subject_name, 
                                MAX(s.subject_score) as score, 
                                MAX(t.total_score) as max_score,
                                MAX(s.expectation) as expectation
                            FROM subject b 
                            JOIN score s ON b.subject_id=s.subject_id
                            JOIN total t ON t.subject_id=b.subject_id AND t.term_id=s.term_id
                            WHERE s.learner_id=%s AND s.grade_id=%s AND s.term_id=%s
                            GROUP BY b.subject_id, b.subject_name
                            ORDER BY b.subject_id""",
                            (learner_id, grade_id, term_id))
                    
                    subjects = cur.fetchall()
                    total_score = 0
                    
                    if not subjects:
                        pdf.set_font("Arial", size=9)
                        pdf.cell(sum(col_widths), 8, "No subject scores available", 1, 1, 'C')
                    else:
                        pdf.set_font("Arial", size=9)
                        for subject_id, name, score, max_score, exp in subjects:
                            total_score += int(score) if str(score).isdigit() else 0
                            
                            name = str(name) if name else "Unknown"
                            score = str(score) if score else "0"
                            exp = str(exp) if exp else ""
                            
                            comment = {
                                "EE": "Exceeding Expectations",
                                "ME": "Meeting Expectations",
                                "AE": "Approaching Expectations"
                            }.get(exp, "Below Expectations")
                            
                            pdf.cell(col_widths[0], 8, name.upper(), 1)
                            pdf.cell(col_widths[1], 8, score, 1, 0, 'C')
                            pdf.cell(col_widths[2], 8, str(subject_out_of), 1, 0, 'C')
                            pdf.cell(col_widths[3], 8, comment, 1, 1, 'L')
                except Exception as query_error:
                    pdf.set_font("Arial", size=9)
                    pdf.cell(sum(col_widths), 8, f"Error loading subjects: {str(query_error)}", 1, 1, 'C')
                    total_score = 0
                
                # Grand Total Row
                pdf.set_font("Arial", 'B', 9)
                pdf.cell(col_widths[0], 8, "GRAND TOTAL", 1)
                pdf.cell(col_widths[1], 8, str(total_score), 1, 0, 'C')
                pdf.cell(col_widths[2], 8, str(grand_total_out_of), 1, 0, 'C')
                
                # Expectations calculation with validation
                try:
                    total_score_int = int(total_score)
                    overall_exp = (
                        "Below Expectations" if total_score_int <= int(grand_total_out_of*0.3) else
                        "Approaching Expectations" if total_score_int <= int(grand_total_out_of*0.5) else
                        "Meeting Expectations" if total_score_int <= int(grand_total_out_of*0.8) else
                        "Exceeding Expectations"
                    )
                except:
                    overall_exp = "Evaluation Not Available"
                
                pdf.cell(col_widths[3], 8, overall_exp, 1, 1, 'L')
                pdf.ln(8)
                
                # Comments Section with fallbacks
                comments = {
                    "Below Expectations": "Add more efforts in your academics, you can do better",
                    "Approaching Expectations": "You can meet expectation. Add more efforts",
                    "Meeting Expectations": "Good Work, you meet expectations well",
                    "Exceeding Expectations": "Excellent, Keep the fire burning"
                }.get(overall_exp, "No comment available")
                
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(90, 8, "CLASS TEACHER'S COMMENT:", 0, 0)
                pdf.cell(0, 8, "OFFICIAL STAMP", 0, 1)
                pdf.set_font("Arial", size=9)
                pdf.multi_cell(90, 8, comments, 0, 'L')
                
                stamp_x, stamp_y = 110, pdf.get_y() - 16
                pdf.rect(stamp_x, stamp_y, 60, 25)
                pdf.ln(5)
                
                # Headteacher comments with fallbacks
                head_comments = {
                    "Below Expectations": "You can meet expectation, add efforts in your academic",
                    "Approaching Expectations": "A fair performance, keep working hard and smart",
                    "Meeting Expectations": "Good performance, you have the potential to exceed expectations",
                    "Exceeding Expectations": "Keep up the good work, your performance shines like a star"
                }.get(overall_exp, "No comment available")
                
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(90, 8, "HEADTEACHER'S COMMENT:", 0, 1)
                pdf.set_font("Arial", size=9)
                pdf.multi_cell(90, 8, head_comments, 0, 'L')
                pdf.ln(5)
                
                # Parent Comments Section
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 8, "PARENT/GUARDIAN'S COMMENTS:", 0, 1)
                pdf.set_font("Arial", size=9)
                line_text = "_" * int((pdf.w - pdf.l_margin - pdf.r_margin) / 2.5)
                for _ in range(3):
                    pdf.cell(0, 8, line_text, 0, 1)
                pdf.ln(3)
                
                # Footer Section
                pdf.set_font("Arial", size=9)
                pdf.cell(90, 5, f"CLOSING DATE: {str(school_info.get('closing_date', ''))}", 0, 0, 'L')
                pdf.cell(0, 5, f"OPENING DATE: {str(school_info.get('opening_date', ''))}", 0, 1, 'R')
                pdf.cell(0, 5, f"Report Sheet Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 0, 'C')
            
            except Exception as page_error:
                QMessageBox.warning(None, "Page Error", f"Error generating page for learner {learner_id}: {str(page_error)}")
                continue  # Skip this learner but continue with others

        # 10. Save with file dialog and proper validation
        default_filename = f"Grade_{grade_name}_Report_Sheets_Term_{selected_term}.pdf"
        default_dir = "D:/TONNIEGIFTED/Documents/programs/AssessmentBoy"
        
        file_dialog = QFileDialog()
        file_dialog.setDefaultSuffix("pdf")
        file_dialog.setNameFilter("PDF Files (*.pdf)")
        file_dialog.setAcceptMode(QFileDialog.AcceptSave)
        file_dialog.setDirectory(default_dir)
        file_dialog.selectFile(default_filename)
        
        if file_dialog.exec_():
            save_path = file_dialog.selectedFiles()[0]
            
            if not save_path:
                QMessageBox.warning(None, "Warning", "No save path selected")
                return
            
            # Ensure PDF extension
            if not save_path.lower().endswith('.pdf'):
                save_path += '.pdf'
            
            # Check if file exists and prompt for overwrite
            if os.path.exists(save_path):
                reply = QMessageBox.question(
                    None,
                    "File Exists", 
                    f"'{os.path.basename(save_path)}' already exists.\nOverwrite?", 
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return  # User chose not to overwrite
            
            try:
                pdf.output(save_path)
                QMessageBox.information(
                    None,
                    "Success",
                    f"Assessment Reports successfully"
                )
            except PermissionError:
                QMessageBox.critical(
                    None,
                    "Error",
                    "Permission denied. Please close the file if it's open elsewhere."
                )
            except Exception as save_error:
                QMessageBox.critical(
                    None,
                    "Error",
                    f"Failed to save file: {str(save_error)}"
                )

    except Exception as e:
        QMessageBox.critical(None, "Error", f"An error occurred:\n{str(e)}")
    finally:
        # Clean up database resources
        try:
            if 'cur' in locals():
                cur.close()
            if 'db' in locals():
                db.close()
        except:
            pass  # Ignore any errors during cleanup

#     # Step 4: Launch app
window = QApplication(sys.argv)
screen = mainwin()
widget = QtWidgets.QStackedWidget()
widget.addWidget(screen)

# Make sure your setFixedSize works correctly with High DPI
widget.setFixedSize(480,375)
widget.setWindowTitle("AssessmentBoy")
widget.show()
sys.exit(window.exec_())
